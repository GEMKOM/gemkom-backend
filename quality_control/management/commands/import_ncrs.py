"""
Management command to import legacy NCRs from an Excel file (.xls or .xlsx).

Usage:
    python manage.py import_ncrs path/to/file.xls [--detected-by <username>] [--dry-run]

Expected columns (case-insensitive, extra columns are ignored):
    Title       — NCR title
    Description — NCR description
    Job No      — JobOrder.job_no (e.g. "280-03-01"); if not found, linked to legacy archive job
    Date        — NCR created_at date (parsed flexibly)

All imported NCRs are created with:
    status      = 'closed'   (they are historical)
    severity    = 'minor'    (default; edit manually if needed)
    defect_type = 'other'    (default; edit manually if needed)
    disposition = 'pending'  (default)
    ncr_number  = auto-generated (NCR-{year}-{seq})
"""

from datetime import datetime

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from projects.models import JobOrder
from quality_control.models import NCR


LEGACY_JOB_NO = "LEGACY-ARCHIVE"
LEGACY_JOB_TITLE = "Legacy Archive (imported NCRs)"


def _get_or_create_legacy_job(created_by: User) -> JobOrder:
    from projects.models import Customer
    customer, _ = Customer.objects.get_or_create(
        code="LEGACY",
        defaults={"name": "Legacy (imported data)"},
    )
    job, _ = JobOrder.objects.get_or_create(
        job_no=LEGACY_JOB_NO,
        defaults={"title": LEGACY_JOB_TITLE, "created_by": created_by, "customer": customer},
    )
    return job


def _parse_date(value):
    if value is None:
        return timezone.now()
    if isinstance(value, datetime):
        return timezone.make_aware(value) if timezone.is_naive(value) else value
    if hasattr(value, "year"):  # date object
        return timezone.make_aware(datetime(value.year, value.month, value.day))
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return timezone.make_aware(datetime.strptime(str(value).strip(), fmt))
        except ValueError:
            continue
    return timezone.now()


def _load_rows_xls(path, sheet_arg):
    """Load rows from a .xls file using xlrd. Returns list of lists of values."""
    try:
        import xlrd
    except ImportError:
        raise CommandError("xlrd is required for .xls files: pip install xlrd")

    try:
        wb = xlrd.open_workbook(path)
    except FileNotFoundError:
        raise CommandError(f"File not found: {path}")
    except Exception as exc:
        raise CommandError(f"Could not open .xls file: {exc}")

    try:
        idx = int(sheet_arg)
        ws = wb.sheet_by_index(idx)
    except (ValueError, TypeError):
        try:
            ws = wb.sheet_by_name(str(sheet_arg))
        except xlrd.biffh.XLRDError:
            raise CommandError(f"Sheet '{sheet_arg}' not found.")
    except IndexError:
        raise CommandError(f"Sheet index {sheet_arg} out of range.")

    rows = []
    for rx in range(ws.nrows):
        row_values = []
        for cx in range(ws.ncols):
            cell = ws.cell(rx, cx)
            # xlrd type 3 = date serial
            if cell.ctype == 3:
                try:
                    dt_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                    row_values.append(datetime(*dt_tuple) if dt_tuple[0] else None)
                except Exception:
                    row_values.append(cell.value)
            else:
                row_values.append(cell.value)
        rows.append(row_values)
    return rows


def _load_rows_xlsx(path, sheet_arg):
    """Load rows from a .xlsx file using openpyxl. Returns list of lists of values."""
    try:
        import openpyxl
    except ImportError:
        raise CommandError("openpyxl is required for .xlsx files: pip install openpyxl")

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError:
        raise CommandError(f"File not found: {path}")
    except Exception as exc:
        raise CommandError(f"Could not open .xlsx file: {exc}")

    try:
        idx = int(sheet_arg)
        ws = wb.worksheets[idx]
    except (ValueError, TypeError):
        if sheet_arg not in wb.sheetnames:
            raise CommandError(f"Sheet '{sheet_arg}' not found.")
        ws = wb[sheet_arg]
    except IndexError:
        raise CommandError(f"Sheet index {sheet_arg} out of range.")

    return [[cell.value for cell in row] for row in ws.iter_rows(values_only=False)]


class Command(BaseCommand):
    help = "Import legacy NCRs from an Excel (.xls or .xlsx) file"

    def add_arguments(self, parser):
        parser.add_argument("excel_file", help="Path to the Excel file (.xls or .xlsx)")
        parser.add_argument(
            "--detected-by",
            default=None,
            help="Username for detected_by / created_by (defaults to first superuser)",
        )
        parser.add_argument(
            "--sheet",
            default=0,
            help="Sheet name or 0-based index (default: first sheet)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database",
        )

    def handle(self, *_args, **options):
        path = options["excel_file"]
        dry_run = options["dry_run"]
        sheet_arg = options["sheet"]

        detected_by = self._resolve_user(options["detected_by"])
        self.stdout.write(f"Using user: {detected_by.username}")

        # Load rows based on extension
        if path.lower().endswith(".xls"):
            rows = _load_rows_xls(path, sheet_arg)
        else:
            rows = _load_rows_xlsx(path, sheet_arg)

        if not rows:
            raise CommandError("The sheet is empty.")

        # Build header map: normalized_name -> col_index
        header_row = rows[0]
        col = {}
        for idx, val in enumerate(header_row):
            normalized = (str(val) if val is not None else "").strip().lower().replace(" ", "_")
            col[normalized] = idx

        self.stdout.write(f"Detected columns: {list(col.keys())}")

        for required in ("title", "description"):
            if required not in col:
                raise CommandError(f"Missing required column: '{required}'. Found: {list(col.keys())}")

        data_rows = rows[1:]
        self.stdout.write(f"Total data rows: {len(data_rows)}")

        created_count = 0
        skipped_count = 0
        error_count = 0
        legacy_job = None

        for row_num, row in enumerate(data_rows, start=2):
            def cell(name):
                idx = col.get(name)
                if idx is None or idx >= len(row):
                    return None
                v = row[idx]
                return str(v).strip() if v is not None else None

            title = cell("title")
            description = cell("description")
            job_no_raw = cell("job_no")
            date_raw = row[col["date"]] if "date" in col and col["date"] < len(row) else None
            corrective_action = cell("kisa_vadede_cozum") or ""
            root_cause = cell("uzun_vadede_cozum") or ""

            if not title and not description:
                skipped_count += 1
                continue

            title = title or "(no title)"
            description = description or ""
            created_at = _parse_date(date_raw)

            # Resolve job order
            job_order = None
            job_note = ""
            if job_no_raw:
                try:
                    job_order = JobOrder.objects.get(job_no=job_no_raw)
                    job_note = f"linked to {job_no_raw}"
                except JobOrder.DoesNotExist:
                    job_note = f"{job_no_raw} not found → legacy archive"

            if job_order is None:
                job_note = job_note or "no job no → legacy archive"
                if not dry_run:
                    if legacy_job is None:
                        legacy_job = _get_or_create_legacy_job(detected_by)
                    job_order = legacy_job

            self.stdout.write(f"  Row {row_num}: \"{title[:60]}\" | {job_note}")

            if dry_run:
                created_count += 1
                continue

            try:
                with transaction.atomic():
                    ncr = NCR(
                        job_order=job_order,
                        title=title,
                        description=description,
                        corrective_action=corrective_action,
                        root_cause=root_cause,
                        defect_type="other",
                        severity="minor",
                        disposition="pending",
                        status="closed",
                        detected_by=detected_by,
                        created_by=detected_by,
                        affected_quantity=1,
                    )
                    ncr.save()
                    # Backfill the original date (auto_now_add bypassed via update)
                    NCR.objects.filter(pk=ncr.pk).update(created_at=created_at)
                    created_count += 1
            except Exception as exc:
                self.stderr.write(f"  ERROR on row {row_num}: {exc}")
                error_count += 1

        self.stdout.write("")
        if dry_run:
            self.stdout.write(self.style.WARNING(
                f"DRY RUN — no changes written. Would create: {created_count}, skipped: {skipped_count}"
            ))
        else:
            self.stdout.write(self.style.SUCCESS(
                f"Done. Created: {created_count}, skipped: {skipped_count}, errors: {error_count}"
            ))
            if legacy_job:
                self.stdout.write(f"Legacy archive job: {LEGACY_JOB_NO} (pk={legacy_job.pk})")

    def _resolve_user(self, username: str | None) -> User:
        if username:
            try:
                return User.objects.get(username=username)
            except User.DoesNotExist:
                raise CommandError(f"User '{username}' not found.")
        user = User.objects.filter(is_superuser=True).order_by("pk").first()
        if user is None:
            user = User.objects.order_by("pk").first()
        if user is None:
            raise CommandError("No users exist in the database. Use --detected-by.")
        return user
